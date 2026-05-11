import sqlite3
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# Load & clean 
print("Loading data...")
df = pd.read_csv('SaaS-Sales.csv')
print(f"  Raw rows: {len(df):,}")

df['Order Date'] = pd.to_datetime(df['Order Date'])
df = df.dropna(subset=['Sales', 'Profit', 'Order Date'])
df = df[df['Sales'] > 0]
df['Margin'] = df['Profit'] / df['Sales']
df['Year']   = df['Order Date'].dt.year
df['Month']  = df['Order Date'].dt.to_period('M')
df['Week']   = df['Order Date'].dt.to_period('W')

print(f"  Clean rows:       {len(df):,}")
print(f"  Date range:       {df['Order Date'].min().date()} → {df['Order Date'].max().date()}")
print(f"  Unique customers: {df['Customer'].nunique():,}")
print(f"  Products:         {df['Product'].nunique():,}")
print(f"  Countries:        {df['Country'].nunique():,}")
print(f"  Total revenue:    ${df['Sales'].sum():,.2f}")
print(f"  Total profit:     ${df['Profit'].sum():,.2f}")
print(f"  Avg margin:       {df['Margin'].mean():.1%}\n")

# Load into SQLite for SQL analysis 
print("Running SQL queries...")
con = sqlite3.connect(':memory:')
df_sql = df.drop(columns=['Month', 'Week'])
df_sql.to_sql('sales', con, index=False, if_exists='replace')

# Weekly revenue with WoW growth (CTE + window via pandas)
weekly_df = pd.read_sql_query("""
    SELECT
        strftime('%Y-W%W', "Order Date")        AS week,
        ROUND(SUM(Sales), 2)                    AS revenue,
        ROUND(SUM(Profit), 2)                   AS profit,
        COUNT(DISTINCT "Order ID")              AS orders,
        COUNT(DISTINCT Customer)                AS customers,
        ROUND(AVG(Discount), 3)                 AS avg_discount
    FROM sales
    GROUP BY 1
    ORDER BY 1
""", con)
weekly_df['wow_growth_pct']    = weekly_df['revenue'].pct_change() * 100
weekly_df['cumulative_revenue'] = weekly_df['revenue'].cumsum()
weekly_df['margin_pct']        = (weekly_df['profit'] / weekly_df['revenue'] * 100).round(1)
weekly_df = weekly_df.round(2)

# Revenue & profit by region
region_df = pd.read_sql_query("""
    WITH region_stats AS (
        SELECT
            Region,
            ROUND(SUM(Sales), 2)            AS revenue,
            ROUND(SUM(Profit), 2)           AS profit,
            COUNT(DISTINCT "Order ID")      AS orders,
            COUNT(DISTINCT Customer)        AS customers
        FROM sales
        GROUP BY Region
    )
    SELECT
        *,
        ROUND(100.0 * profit / revenue, 1) AS margin_pct,
        RANK() OVER (ORDER BY revenue DESC) AS revenue_rank
    FROM region_stats
    ORDER BY revenue DESC
""", con)

print("── Revenue by region ──")
print(region_df.to_string(index=False))

# Product performance
product_df = pd.read_sql_query("""
    SELECT
        Product,
        ROUND(SUM(Sales), 2)            AS revenue,
        ROUND(SUM(Profit), 2)           AS profit,
        COUNT(DISTINCT "Order ID")      AS orders,
        ROUND(AVG(Discount) * 100, 1)   AS avg_discount_pct,
        ROUND(100.0 * SUM(Profit) / SUM(Sales), 1) AS margin_pct
    FROM sales
    GROUP BY Product
    ORDER BY revenue DESC
""", con)

print("\n── Product performance ──")
print(product_df.to_string(index=False))

# Customer segment analysis
segment_df = pd.read_sql_query("""
    SELECT
        Segment,
        ROUND(SUM(Sales), 2)            AS revenue,
        ROUND(SUM(Profit), 2)           AS profit,
        COUNT(DISTINCT Customer)        AS customers,
        COUNT(DISTINCT "Order ID")      AS orders,
        ROUND(SUM(Sales) / COUNT(DISTINCT Customer), 2) AS revenue_per_customer,
        ROUND(100.0 * SUM(Profit) / SUM(Sales), 1)      AS margin_pct
    FROM sales
    GROUP BY Segment
    ORDER BY revenue DESC
""", con)

print("\n── Segment performance ──")
print(segment_df.to_string(index=False))

# Discount impact — binned analysis
discount_df = pd.read_sql_query("""
    SELECT
        CASE
            WHEN Discount = 0         THEN '0% (no discount)'
            WHEN Discount <= 0.10     THEN '1–10%'
            WHEN Discount <= 0.20     THEN '11–20%'
            WHEN Discount <= 0.30     THEN '21–30%'
            ELSE '30%+'
        END                                             AS discount_band,
        COUNT(*)                                        AS orders,
        ROUND(SUM(Sales), 2)                            AS revenue,
        ROUND(AVG(100.0 * Profit / Sales), 1)           AS avg_margin_pct
    FROM sales
    GROUP BY 1
    ORDER BY 2
""", con)

print("\n── Discount impact on margin ──")
print(discount_df.to_string(index=False))

# Top 10 customers by revenue (window function)
top_customers = pd.read_sql_query("""
    SELECT
        Customer,
        Industry,
        Segment,
        Region,
        ROUND(SUM(Sales), 2)                            AS revenue,
        ROUND(SUM(Profit), 2)                           AS profit,
        COUNT(DISTINCT "Order ID")                      AS orders,
        ROUND(100.0 * SUM(Profit) / SUM(Sales), 1)      AS margin_pct,
        RANK() OVER (ORDER BY SUM(Sales) DESC)          AS rank
    FROM sales
    GROUP BY Customer, Industry, Segment, Region
    ORDER BY revenue DESC
    LIMIT 15
""", con)

print("\n── Top 15 customers ──")
print(top_customers.to_string(index=False))

# Charts 
print("\nGenerating charts...")
BLUE   = '#1D4ED8'
GREEN  = '#15803D'
RED    = '#DC2626'
COLORS = ['#1D4ED8','#2563EB','#3B82F6','#60A5FA','#93C5FD','#BFDBFE','#1E40AF','#172554']

fig, axes = plt.subplots(2, 3, figsize=(18, 11))
fig.suptitle('SaaS Sales Reporting Dashboard\nAWS SaaS Sales Dataset · Kaggle',
             fontsize=15, fontweight='bold', y=1.01)

# Monthly revenue (aggregate weekly to monthly for cleaner chart)
monthly = df.groupby(df['Order Date'].dt.to_period('M')).agg(
    revenue=('Sales','sum'), profit=('Profit','sum')).reset_index()
monthly['date'] = monthly['Order Date'].dt.to_timestamp()

ax = axes[0,0]
ax.fill_between(monthly['date'], monthly['revenue'], alpha=0.12, color=BLUE)
ax.plot(monthly['date'], monthly['revenue'], color=BLUE, linewidth=2.5, marker='o', markersize=4, label='Revenue')
ax.plot(monthly['date'], monthly['profit'],  color=GREEN, linewidth=2, linestyle='--', marker='s', markersize=3, label='Profit')
ax.set_title('Monthly Revenue & Profit ($)', fontweight='bold')
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x,_: f'${x/1000:.0f}k'))
ax.tick_params(axis='x', rotation=30)
ax.legend(fontsize=9)
ax.grid(axis='y', alpha=0.3)

# Revenue by region
ax = axes[0,1]
bars = ax.bar(region_df['Region'], region_df['revenue'], color=COLORS[:len(region_df)], alpha=0.88)
ax.set_title('Revenue by Region', fontweight='bold')
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x,_: f'${x/1000:.0f}k'))
ax.tick_params(axis='x', rotation=20)
for bar, val in zip(bars, region_df['revenue']):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+500,
            f'${val/1000:.0f}k', ha='center', fontsize=8)
ax.grid(axis='y', alpha=0.3)

# Product revenue + margin
ax = axes[0,2]
x = np.arange(len(product_df))
bars = ax.bar(x, product_df['revenue'], color=BLUE, alpha=0.85, label='Revenue')
ax.set_xticks(x)
ax.set_xticklabels(product_df['Product'], rotation=25, ha='right', fontsize=8)
ax.set_title('Revenue & Margin by Product', fontweight='bold')
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x,_: f'${x/1000:.0f}k'))
ax2 = ax.twinx()
ax2.plot(x, product_df['margin_pct'], color=RED, marker='D', linewidth=2, markersize=6, label='Margin %')
ax2.set_ylabel('Margin %', color=RED)
ax2.tick_params(axis='y', labelcolor=RED)
ax.grid(axis='y', alpha=0.3)

# Segment performance
ax = axes[1,0]
width = 0.35
x = np.arange(len(segment_df))
ax.bar(x - width/2, segment_df['revenue'], width, color=BLUE,  alpha=0.85, label='Revenue')
ax.bar(x + width/2, segment_df['profit'],  width, color=GREEN, alpha=0.85, label='Profit')
ax.set_xticks(x)
ax.set_xticklabels(segment_df['Segment'], rotation=15)
ax.set_title('Revenue & Profit by Segment', fontweight='bold')
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x,_: f'${x/1000:.0f}k'))
ax.legend(fontsize=9)
ax.grid(axis='y', alpha=0.3)

# Discount impact on margin
ax = axes[1,1]
colors_disc = [GREEN if m > 0 else RED for m in discount_df['avg_margin_pct']]
bars = ax.bar(discount_df['discount_band'], discount_df['avg_margin_pct'],
              color=colors_disc, alpha=0.85)
ax.set_title('Avg Profit Margin by Discount Band', fontweight='bold')
ax.set_ylabel('Avg Margin %')
ax.tick_params(axis='x', rotation=15)
ax.axhline(0, color='black', linewidth=0.8)
for bar, val in zip(bars, discount_df['avg_margin_pct']):
    ax.text(bar.get_x()+bar.get_width()/2,
            bar.get_height() + (1 if val >= 0 else -3),
            f'{val:.1f}%', ha='center', fontsize=9, fontweight='bold')
ax.grid(axis='y', alpha=0.3)

# Cumulative revenue
ax = axes[1,2]
ax.plot(range(len(weekly_df)), weekly_df['cumulative_revenue'],
        color=BLUE, linewidth=2.5)
ax.fill_between(range(len(weekly_df)), weekly_df['cumulative_revenue'],
                alpha=0.12, color=BLUE)
ax.set_title('Cumulative Revenue Over Time', fontweight='bold')
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x,_: f'${x/1e6:.1f}M'))
ax.set_xlabel('Week')
ax.grid(axis='y', alpha=0.3)

plt.tight_layout()
plt.savefig('saas_analysis.png', dpi=150, bbox_inches='tight')
print("Saved saas_analysis.png")

# Excel Report
print("Building Excel report...")

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
STRIPE_FILL  = PatternFill("solid", fgColor="EBF3FB")
TOTAL_FILL   = PatternFill("solid", fgColor="D4E6F5")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
thin = Side(style='thin', color="C0C0C0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_header(ws, row, cols, col_start=1):
    for c, label in enumerate(cols, col_start):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def write_row(ws, row_num, values, col_start=1, stripe=False, bold=False):
    fill = STRIPE_FILL if stripe else WHITE_FILL
    for c, val in enumerate(values, col_start):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.fill = fill; cell.border = BORDER
        cell.font = BOLD_FONT if bold else BODY_FONT
        cell.alignment = Alignment(
            horizontal='right' if isinstance(val, (int, float)) else 'left')

def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

wb = openpyxl.Workbook()

# Cover
ws = wb.active; ws.title = "Cover"
ws.sheet_view.showGridLines = False
ws['B2'] = "SaaS Sales Reporting Dashboard"
ws['B2'].font = Font(name="Calibri", bold=True, size=22, color="1E3A5F")
ws['B3'] = "AWS SaaS Sales Dataset · Product Analytics"
ws['B3'].font = Font(name="Calibri", size=13, color="555555")
ws['B5'] = "Sheets"
ws['B5'].font = BOLD_FONT
for i, (name, desc) in enumerate([
    ("Weekly Revenue",    "Week-by-week sales, WoW growth, cumulative total, margin"),
    ("Region Analysis",   "Revenue, profit, and margin by geographic region"),
    ("Product Analysis",  "Revenue, profit, and discount impact per product"),
    ("Segment Analysis",  "Performance across SMB, Mid-Market, Enterprise segments"),
    ("Discount Impact",   "How discounting affects profit margins"),
    ("Top Customers",     "Top 15 customers by lifetime revenue"),
], 7):
    ws.cell(i, 2, f"• {name}").font = BOLD_FONT
    ws.cell(i, 3, desc).font = BODY_FONT
ws['B14'] = f"Source: https://www.kaggle.com/datasets/nnthanh101/aws-saas-sales"
ws['B14'].font = Font(name="Calibri", size=9, color="888888", italic=True)
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 52

# Weekly Revenue
ws2 = wb.create_sheet("Weekly Revenue")
ws2.sheet_view.showGridLines = False
ws2['A1'] = "Weekly Revenue Trend"; ws2['A1'].font = TITLE_FONT
cols = ["Week","Revenue ($)","Profit ($)","Orders","Customers","Avg Discount","Margin %","WoW Growth %","Cumulative ($)"]
write_header(ws2, 3, cols)
for i, row in weekly_df.iterrows():
    rn = i + 4
    wow = row['wow_growth_pct'] if not pd.isna(row['wow_growth_pct']) else ''
    vals = [row['week'], row['revenue'], row['profit'], row['orders'],
            row['customers'], row['avg_discount'], row['margin_pct'], wow, row['cumulative_revenue']]
    write_row(ws2, rn, vals, stripe=i%2==1)
    for col in [2,3,9]:
        ws2.cell(rn,col).number_format = '#,##0.00'
    ws2.cell(rn,7).number_format = '0.0"%"'
    if wow != '':
        ws2.cell(rn,8).number_format = '+0.0;-0.0;0.0'
set_widths(ws2, [12,14,14,9,11,13,10,14,16])

chart = LineChart()
chart.title = "Weekly Revenue ($)"; chart.style = 10
chart.height = 12; chart.width = 22
data = Reference(ws2, min_col=2, min_row=3, max_row=3+len(weekly_df))
chart.add_data(data, titles_from_data=True)
cats = Reference(ws2, min_col=1, min_row=4, max_row=3+len(weekly_df))
chart.set_categories(cats)
chart.series[0].graphicalProperties.line.solidFill = "1D4ED8"
chart.series[0].graphicalProperties.line.width = 18000
ws2.add_chart(chart, "K3")

# Region
ws3 = wb.create_sheet("Region Analysis")
ws3.sheet_view.showGridLines = False
ws3['A1'] = "Revenue & Profit by Region"; ws3['A1'].font = TITLE_FONT
write_header(ws3, 3, ["Region","Revenue ($)","Profit ($)","Orders","Customers","Margin %","Rank"])
for i, row in region_df.iterrows():
    rn = i+4
    write_row(ws3, rn, [row['Region'],row['revenue'],row['profit'],
                         row['orders'],row['customers'],row['margin_pct'],row['revenue_rank']], stripe=i%2==1)
    ws3.cell(rn,2).number_format = '#,##0.00'
    ws3.cell(rn,3).number_format = '#,##0.00'
    ws3.cell(rn,6).number_format = '0.0"%"'
set_widths(ws3, [16,14,14,9,11,10,7])

# Product
ws4 = wb.create_sheet("Product Analysis")
ws4.sheet_view.showGridLines = False
ws4['A1'] = "Product Performance"; ws4['A1'].font = TITLE_FONT
write_header(ws4, 3, ["Product","Revenue ($)","Profit ($)","Orders","Avg Discount %","Margin %"])
for i, row in product_df.iterrows():
    rn = i+4
    write_row(ws4, rn, [row['Product'],row['revenue'],row['profit'],
                         row['orders'],row['avg_discount_pct'],row['margin_pct']], stripe=i%2==1)
    ws4.cell(rn,2).number_format = '#,##0.00'
    ws4.cell(rn,3).number_format = '#,##0.00'
set_widths(ws4, [22,14,14,9,16,10])

# Segment
ws5 = wb.create_sheet("Segment Analysis")
ws5.sheet_view.showGridLines = False
ws5['A1'] = "Customer Segment Performance"; ws5['A1'].font = TITLE_FONT
write_header(ws5, 3, ["Segment","Revenue ($)","Profit ($)","Customers","Orders","Rev / Customer ($)","Margin %"])
for i, row in segment_df.iterrows():
    rn = i+4
    write_row(ws5, rn, [row['Segment'],row['revenue'],row['profit'],row['customers'],
                         row['orders'],row['revenue_per_customer'],row['margin_pct']], stripe=i%2==1)
    for col in [2,3,6]: ws5.cell(rn,col).number_format = '#,##0.00'
set_widths(ws5, [18,14,14,11,9,20,10])

# Discount impact
ws6 = wb.create_sheet("Discount Impact")
ws6.sheet_view.showGridLines = False
ws6['A1'] = "Discount Impact on Profit Margin"; ws6['A1'].font = TITLE_FONT
write_header(ws6, 3, ["Discount Band","Orders","Revenue ($)","Avg Margin %"])
for i, row in discount_df.iterrows():
    rn = i+4
    write_row(ws6, rn, [row['discount_band'],row['orders'],row['revenue'],row['avg_margin_pct']], stripe=i%2==1)
    ws6.cell(rn,3).number_format = '#,##0.00'
set_widths(ws6, [20,9,14,14])

# Top customers
ws7 = wb.create_sheet("Top Customers")
ws7.sheet_view.showGridLines = False
ws7['A1'] = "Top 15 Customers by Revenue"; ws7['A1'].font = TITLE_FONT
write_header(ws7, 3, ["Rank","Customer","Industry","Segment","Region","Revenue ($)","Profit ($)","Orders","Margin %"])
for i, row in top_customers.iterrows():
    rn = i+4
    write_row(ws7, rn, [row['rank'],row['Customer'],row['Industry'],row['Segment'],
                         row['Region'],row['revenue'],row['profit'],row['orders'],row['margin_pct']], stripe=i%2==1)
    for col in [6,7]: ws7.cell(rn,col).number_format = '#,##0.00'
set_widths(ws7, [6,26,18,14,12,14,14,9,10])

wb.save('saas_sales_report.xlsx')
print("Saved saas_sales_report.xlsx")